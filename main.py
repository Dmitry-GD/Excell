import openpyxl as op
import openpyxl.styles.numbers
import os
import time
'''
# ОБРАБОТКА ФАЙЛА ШПАРГАЛКИ И ЗАПОЛНЕНИЕ ШАБЛОНОВ ПО НЕМУ
def read_template(sheet_name: str, template_name: dict[str, list]) -> None:
    """
    Функция считывания полей шаблона
    :param sheet_name: str
    :param template_name: dict[str, list]
    :return: None
    """
    sheet = wb[sheet_name]
    max_row = sheet.max_row
    if sheet.cell(row=7, column=5).value == None:
        for i in range(8, max_row):
            template_name[sheet.cell(row=i, column=5).value.lower().rstrip().replace('\n', '')] = [sheet.cell(row=i, column=9).value,
                                                                        sheet.cell(row=i, column=10).value,
                                                                        sheet.cell(row=i, column=14).value,
                                                                        sheet.cell(row=i, column=15).value,
                                                                        sheet.cell(row=i, column=16).value,
                                                                        sheet.cell(row=i, column=17).value,
                                                                        sheet.cell(row=i, column=20).value
                                                                        ]
            if 'фоив' in sheet_name.lower() and 'государственный кадастровый учет недвижимого имущества' in sheet.cell(row=i, column=5).value.lower().rstrip().replace('\n', ''):
                template_name[sheet.cell(row=i, column=5).value.lower().rstrip().replace('\n', '')].append(sheet.cell(row=i, column=7).value)
                template_name[sheet.cell(row=i, column=5).value.lower().rstrip().replace('\n', '')].append(sheet.cell(row=i, column=8).value)
                template_name[sheet.cell(row=i, column=5).value.lower().rstrip().replace('\n', '')].append(sheet.cell(row=i, column=12).value)
                template_name[sheet.cell(row=i, column=5).value.lower().rstrip().replace('\n', '')].append(sheet.cell(row=i, column=13).value)
            if 'фоив' in sheet_name.lower() and 'предоставление сведений, содержащихся в едином государственном реестре' in sheet.cell(row=i, column=5).value.lower().rstrip().replace('\n', ''):
                template_name[sheet.cell(row=i, column=5).value.lower().rstrip().replace('\n', '')].append(sheet.cell(row=i, column=7).value)
                template_name[sheet.cell(row=i, column=5).value.lower().rstrip().replace('\n', '')].append(sheet.cell(row=i, column=12).value)

def write_template(sheet_name: str, template_name: dict[str, list], not_tosp = True) -> None:
    """
    Функция записи полей таблицы из шаблона
    :param sheet_name: str
        :param template_name: dict[str, list]
        :return: None
        """
    sheet = wb[sheet_name]
    max_row = sheet.max_row
    key_list = tuple(template_name.keys())
    change_key_counter = 0
    errors_name = []
    if sheet.cell(row=7, column=5).value == None:
        for i in range(8, max_row):
            if sheet.cell(row=i, column=5).value.lower().rstrip().replace('\n', '') == key_list[i-8]:
                key = sheet.cell(row=i, column=5).value.lower().rstrip().replace('\n', '')
            else:
                key = key_list[i-8]
                change_key_counter += 1
                temp = tuple([i-7, sheet.cell(row=i, column=5).value.lower().rstrip().replace('\n', ''), key_list[i-8]])
                errors_name.append(temp)
            sheet.cell(row=i, column=9).value = template_name[key][0]
            sheet.cell(row=i, column=10).value = template_name[key][1]
            sheet.cell(row=i, column=14).value = template_name[key][2]
            sheet.cell(row=i, column=15).value = template_name[key][3]
            sheet.cell(row=i, column=16).value = template_name[key][4]
            sheet.cell(row=i, column=17).value = template_name[key][5]
            sheet.cell(row=i, column=20).value = template_name[key][6]
            if not_tosp:
                if 'государственный кадастровый учет недвижимого имущества' in sheet.cell(row=i, column=5).value.lower().rstrip().replace('\n', ''):
                    sheet.cell(row=i, column=7).value = template_name[key][7]
                    sheet.cell(row=i, column=8).value = template_name[key][8]
                    sheet.cell(row=i, column=12).value = template_name[key][9]
                    sheet.cell(row=i, column=13).value = template_name[key][10]
                if 'предоставление сведений, содержащихся в едином государственном реестре' in sheet.cell(row=i, column=5).value.lower().rstrip().replace('\n', ''):
                    sheet.cell(row=i, column=7).value = template_name[key][7]
                    sheet.cell(row=i, column=12).value = template_name[key][8]
            sheet.cell(row=max_row, column=19).value = 0
    if change_key_counter > 0:
        print('*****************************')
        print('В листе', sheet_name, 'было найдено', change_key_counter, 'несостыковок названий услуг:')
        for item in errors_name:
            print(f'   В услуге № {item[0]}:\nНазвание в файле:     {item[1]}\nНазвание в шпаргалке: {item[2]}')
            print('--------------------------------')
        print('*****************************')

# Формирование списка файлв для обработки
file_list = os.listdir()    # просмотр файловой системы
template_file_name = ['', '', '', '', '', '']
for item in file_list:
    if 'шпаргалка' in item.lower():
        template_file = item      # Эксель файл с шаблоном заполнения
    elif 'фоив' in item.lower() and 'тосп' not in item.lower():
        template_file_name[0] = item
    elif 'роив' in item.lower():
        template_file_name[1] = item
    elif 'омсу' in item.lower():
        template_file_name[2] = item
    elif 'иные' in item.lower() and 'тосп' not in item.lower():
        template_file_name[3] = item
    elif 'фоив' in item.lower() and 'тосп' in item.lower():
        template_file_name[4] = item
    elif 'иные' in item.lower() and 'тосп' in item.lower():
        template_file_name[5] = item
print(f'Найдено {len(template_file_name)} файлов с именами :{template_file_name}')

# Обработка файла ШПАРГАЛКИ
wb = op.load_workbook(template_file, data_only=True)
template_foiv = {}
template_roiv = {}
template_omsu = {}
template_inie = {}
sheet_list = []
for item in wb.sheetnames:
    if 'фоив' in item.lower() and 'бизнес' not in item.lower():
        sheet_list.append(item)
    elif 'роив' in item.lower() and 'бизнес' not in item.lower():
        sheet_list.append(item)
    elif 'омсу' in item.lower() and 'бизнес' not in item.lower():
        sheet_list.append(item)
    elif 'иных' in item.lower() and 'бизнес' not in item.lower():
        sheet_list.append(item)
read_template(sheet_list[0], template_foiv)
read_template(sheet_list[1], template_roiv)
read_template(sheet_list[2], template_omsu)
read_template(sheet_list[3], template_inie)
wb.close()

# НАЧАЛО ЗАПОЛНЕНИЯ ШАБЛОНОВ по шпаргалке

# Открываем ФОИВ для заполнения
wb = op.load_workbook(template_file_name[0], data_only=True)
print(f'Начало обработки файла {template_file_name[0]}')
write_template('5007', template_foiv)
write_template('366', template_foiv)
print(f'Обработка {template_file_name[0]} завершена!!!')
wb.save('ФОИВ заполненный по шпаргалке.xlsx')

# Открываем РОИВ для заполнения
wb = op.load_workbook(template_file_name[1], data_only=True)
print(f'Начало обработки файла {template_file_name[1]}')
write_template('5007', template_roiv)
write_template('366', template_roiv)
print(f'Обработка {template_file_name[1]} завершена!!!')
wb.save('РОИВ заполненный по шпаргалке.xlsx')

# Открываем ОМСУ для заполнения
wb = op.load_workbook(template_file_name[2], data_only=True)
print(f'Начало обработки файла {template_file_name[2]}')
write_template('5007', template_omsu)
write_template('366', template_omsu)
print(f'Обработка {template_file_name[2]} завершена!!!')
wb.save('ОМСУ заполненный по шпаргалке.xlsx')

# Открываем ИНЫЕ УСЛУГИ для заполнения
wb = op.load_workbook(template_file_name[3], data_only=True)
print(f'Начало обработки файла {template_file_name[3]}')
write_template('5007', template_inie)
write_template('366', template_inie)
print(f'Обработка {template_file_name[3]} завершена!!!')
wb.save('ИНЫЕ УСЛУГИ заполненный по шпаргалке.xlsx')

# Открываем ФОИВ ТОСП для заполнения
wb = op.load_workbook(template_file_name[4], data_only=True)
print(f'Начало обработки файла {template_file_name[4]}')
write_template('9000437', template_foiv, not_tosp=False)
write_template('9000436', template_foiv, not_tosp=False)
write_template('9000441', template_foiv, not_tosp=False)
write_template('9000438', template_foiv, not_tosp=False)
write_template('9000440', template_foiv, not_tosp=False)
write_template('9000439', template_foiv, not_tosp=False)
print(f'Обработка {template_file_name[4]} завершена!!!')
wb.save('ФОИВ ТОСП заполненный по шпаргалке.xlsx')

# Открываем ИНЫЕ УСЛУГИ ТОСП для заполнения
wb = op.load_workbook(template_file_name[5], data_only=True)
print(f'Начало обработки файла {template_file_name[5]}')
write_template('9000437', template_inie, not_tosp=False)
write_template('9000436', template_inie, not_tosp=False)
write_template('9000441', template_inie, not_tosp=False)
write_template('9000438', template_inie, not_tosp=False)
write_template('9000440', template_inie, not_tosp=False)
write_template('9000439', template_inie, not_tosp=False)
print(f'Обработка {template_file_name[5]} завершена!!!')
wb.save('ИНЫЕ УСЛУГИ ТОСП заполненный по шпаргалке.xlsx')

all_uslugi = list(template_foiv.keys()) + list(template_roiv.keys()) + list(template_omsu.keys()) + list(template_inie.keys()) # полный список услуг
'''



# !!! !!! !!! Начало обработки отчетов выгруженных из АИС !!! !!! !!!
result_ilicha = {}
result_lenina = {}

# ОТЧЕТЫ ПВД !!!   !!!   !!!   !!!   !!!   !!!   !!!   !!!   !!!   !!!   !!!
# Обработка отчета ПВД по Ильича <<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>>
file_list = os.listdir()    # просмотр файловой системы
for item in file_list:
    if 'ильича' in item.lower() and 'пвд' in item.lower():
        file = item
wb = op.load_workbook(file, data_only=True)
# <<< ЛИСТ ПРИНЯТЫХ ОБРАЩЕНИЙ >>>
sheet = wb['Прием обращений']
max_row = sheet.max_row
temp_predostavlenie = set()
temp_registration = set()
for i in range(2, max_row + 1):
    if 'предоставление' in sheet.cell(row=i, column=5).value.lower():
        temp_predostavlenie.add(sheet.cell(row=i, column=2).value)
    else:
        temp_registration.add(sheet.cell(row=i, column=2).value)
result_ilicha['предоставление сведений, содержащихся в Едином государственном реестре недвижимости'] = [None, None, None, None, None, None, None, None, None]
result_ilicha['государственный кадастровый учет недвижимого имущества и (или) государственная регистрация прав на недвижимое имущество'] = [None, None, None, None, None, None, None, None, None, None, None]
result_ilicha['предоставление сведений, содержащихся в Едином государственном реестре недвижимости'][7] = len(temp_predostavlenie)
result_ilicha['государственный кадастровый учет недвижимого имущества и (или) государственная регистрация прав на недвижимое имущество'][7] = len(temp_registration)
# <<< ЛИСТ ВЫДАННЫХ ОБРАЩЕНИЙ >>>
sheet = wb['Выданные обращения']
max_row = sheet.max_row
temp_predostavlenie = set()
temp_registration = set()
for i in range(2, max_row + 1):
    if 'предоставление' in sheet.cell(row=i, column=6).value.lower():
        temp_predostavlenie.add(sheet.cell(row=i, column=3).value)
    else:
        temp_registration.add(sheet.cell(row=i, column=3).value)
result_ilicha['предоставление сведений, содержащихся в Едином государственном реестре недвижимости'][8] = len(temp_predostavlenie)
result_ilicha['государственный кадастровый учет недвижимого имущества и (или) государственная регистрация прав на недвижимое имущество'][9] = len(temp_registration)
wb.close()
# Обработка отчета ПВД по Ленина <<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>>
file_list = os.listdir()    # просмотр файловой системы
for item in file_list:
    if 'ленина' in item.lower() and 'пвд' in item.lower():
        file = item
wb = op.load_workbook(file, data_only=True)
# <<< ЛИСТ ПРИНЯТЫХ ОБРАЩЕНИЙ >>>
sheet = wb['Прием обращений']
max_row = sheet.max_row
temp_predostavlenie = set()
temp_registration = set()
for i in range(2, max_row + 1):
    if 'предоставление' in sheet.cell(row=i, column=5).value.lower():
        temp_predostavlenie.add(sheet.cell(row=i, column=2).value)
    else:
        temp_registration.add(sheet.cell(row=i, column=2).value)
result_lenina['предоставление сведений, содержащихся в Едином государственном реестре недвижимости'] = [None, None, None, None, None, None, None, None, None]
result_lenina['государственный кадастровый учет недвижимого имущества и (или) государственная регистрация прав на недвижимое имущество'] = [None, None, None, None, None, None, None, None, None, None, None]
result_lenina['предоставление сведений, содержащихся в Едином государственном реестре недвижимости'][7] = len(temp_predostavlenie)
result_lenina['государственный кадастровый учет недвижимого имущества и (или) государственная регистрация прав на недвижимое имущество'][7] = len(temp_registration)
# <<< ЛИСТ ВЫДАННЫХ ОБРАЩЕНИЙ >>>
sheet = wb['Выданные обращения']
max_row = sheet.max_row
temp_predostavlenie = set()
temp_registration = set()
for i in range(2, max_row + 1):
    if 'предоставление' in sheet.cell(row=i, column=6).value.lower():
        temp_predostavlenie.add(sheet.cell(row=i, column=3).value)
    else:
        temp_registration.add(sheet.cell(row=i, column=3).value)
result_lenina['предоставление сведений, содержащихся в Едином государственном реестре недвижимости'][8] = len(temp_predostavlenie)
result_lenina['государственный кадастровый учет недвижимого имущества и (или) государственная регистрация прав на недвижимое имущество'][9] = len(temp_registration)
wb.close()



# ОТЧЕТЫ ПО ВЫДАЧЕ !!!   !!!   !!!   !!!   !!!   !!!   !!!   !!!   !!!   !!!   !!!
# Обработка отчета по выдаче по Ильича <<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>>
file_list = os.listdir()    # просмотр файловой системы
for item in file_list:
    if 'ильича' in item.lower() and 'выдача' in item.lower():
        file = item
wb = op.load_workbook(file, data_only=True)
sheet = wb['Услуги МФЦ']
max_row = sheet.max_row
for i in range(2, max_row + 1):
    if sheet.cell(row=i, column=2).value.lower().rstrip().replace('\n', '') not in result_ilicha:
        result_ilicha[sheet.cell(row=i, column=2).value.lower().rstrip().replace('\n', '')] = [None, None, None, None, None, None, None]
    result_ilicha[sheet.cell(row=i, column=2).value.lower().rstrip().replace('\n', '')][2] = int(sheet.cell(row=i, column=4).value)
    result_ilicha[sheet.cell(row=i, column=2).value.lower().rstrip().replace('\n', '')][3] = int(sheet.cell(row=i, column=4).value)
wb.close()
# Обработка отчета по выдаче по Ленина <<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>>
file_list = os.listdir()    # просмотр файловой системы
for item in file_list:
    if 'ленина' in item.lower() and 'выдача' in item.lower():
        file = item
wb = op.load_workbook(file, data_only=True)
sheet = wb['Услуги МФЦ']
max_row = sheet.max_row
for i in range(2, max_row + 1):
    if sheet.cell(row=i, column=2).value.lower().rstrip().replace('\n', '') not in result_lenina:
        result_lenina[sheet.cell(row=i, column=2).value.lower().rstrip().replace('\n', '')] = [None, None, None, None, None, None, None]
    result_lenina[sheet.cell(row=i, column=2).value.lower().rstrip().replace('\n', '')][2] = int(sheet.cell(row=i, column=4).value)
    result_lenina[sheet.cell(row=i, column=2).value.lower().rstrip().replace('\n', '')][3] = int(sheet.cell(row=i, column=4).value)
wb.close()



# ОТЧЕТЫ ПО ПРИЁМУ !!!   !!!   !!!   !!!   !!!   !!!   !!!   !!!   !!!   !!!   !!!
# Обработка отчета по приёму по Ильича <<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>>
file_list = os.listdir()    # просмотр файловой системы
for item in file_list:
    if 'ильича' in item.lower() and 'приём' in item.lower():
        file = item
wb = op.load_workbook(file, data_only=True)
sheet = wb['Услуги МФЦ']
max_row = sheet.max_row
for i in range(2, max_row + 1):
    if sheet.cell(row=i, column=2).value.lower().rstrip().replace('\n', '') not in result_ilicha:
        result_ilicha[sheet.cell(row=i, column=2).value.lower().rstrip().replace('\n', '')] = [None, None, None, None, None, None, None]
    result_ilicha[sheet.cell(row=i, column=2).value.lower().rstrip().replace('\n', '')][0] = int(sheet.cell(row=i, column=4).value)
wb.close()
# Обработка отчета по приёму по Ленина <<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>>
file_list = os.listdir()    # просмотр файловой системы
for item in file_list:
    if 'ленина' in item.lower() and 'приём' in item.lower():
        file = item
wb = op.load_workbook(file, data_only=True)
sheet = wb['Услуги МФЦ']
max_row = sheet.max_row
for i in range(2, max_row + 1):
    if sheet.cell(row=i, column=2).value.lower().rstrip().replace('\n', '') not in result_lenina:
        result_lenina[sheet.cell(row=i, column=2).value.lower().rstrip().replace('\n', '')] = [None, None, None, None, None, None, None]
    result_lenina[sheet.cell(row=i, column=2).value.lower().rstrip().replace('\n', '')][0] = int(sheet.cell(row=i, column=4).value)
wb.close()



# ОТЧЕТЫ ПО СВЕРКАМ !!!   !!!   !!!   !!!   !!!   !!!   !!!   !!!   !!!   !!!   !!!
# LOOOL пока пропустил, есть вопросы



# ОТЧЕТЫ МИНЭКОНОМРАЗВИТИЯ !!!   !!!   !!!   !!!   !!!   !!!   !!!   !!!   !!!   !!!   !!!
minekonom_result_ilicha = {}
minekonom_result_lenina = {}
sheet_list = ['федеральные услуги', 'роив', 'омсу', 'услуги иных организаций']
def read_mineko(sheet_name: str, result_dict: dict[str, list], consultation_list: list) -> None:
    """
    Считывает переданный лист тамлицы инэконом
    :param sheet_name: str
    :param result_dict: dict[str, list]
    :param consultation_list: list
    :return: None
    """

    def change_value(cell_value: str | float) -> str | int | None:
        """
        Проверяет значение ячейки и выводит в нужном формате
        :param cell_value: str | float
        :return: str | int | None
        """
        if isinstance(cell_value, float):
            return int(cell_value)
        elif cell_value == '':
            return None
        elif 'нет' in cell_value:
            return cell_value
        else:
            return int(cell_value)

    sheet = wb[sheet_name]
    max_row = sheet.max_row
    for i in range(8, max_row):
        key = sheet.cell(row=i, column=5).value.lower().rstrip().replace('\n', '')
        result_dict[key] = [change_value(sheet.cell(row=i, column=9).value),
                            change_value(sheet.cell(row=i, column=10).value),
                            change_value(sheet.cell(row=i, column=14).value),
                            change_value(sheet.cell(row=i, column=15).value),
                            change_value(sheet.cell(row=i, column=16).value),
                            change_value(sheet.cell(row=i, column=17).value),
                            None
                            ]
    consultation_list.append(change_value(sheet.cell(row=max_row, column=19).value))

# Обработка отчета минэконмразвития по Ильича <<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>>
file_list = os.listdir()    # просмотр файловой системы
for item in file_list:
    if 'ильича' in item.lower() and 'минэконом' in item.lower():
        file = item
wb = op.load_workbook(file, data_only=True)
consultation_ilicha = []
for sheet in sheet_list:
    read_mineko(sheet, minekonom_result_ilicha, consultation_ilicha)
wb.close()
# Обработка отчета минэконмразвития по Ленина <<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>><<<>>>
file_list = os.listdir()    # просмотр файловой системы
for item in file_list:
    if 'ленина' in item.lower() and 'минэконом' in item.lower():
        file = item
wb = op.load_workbook(file, data_only=True)
consultation_lenina = []
for sheet in sheet_list:
    read_mineko(sheet, minekonom_result_lenina, consultation_lenina)
wb.close()



'''
# Раздел вывода результатов на экран
print('Результаты Ильича:')
print(*result_ilicha.items(), sep='\n')
print()
print('Результаты Ленина:')
print(*result_lenina.items(), sep='\n')
'''



'''
# Разбор повторяющихся услуг в МИНЭКОНОМ И ОСТАЛЬНЫХ ОТЧЕТАХ
mineko_result_only_positive_uslugi_list_ilicha = list(map(lambda x: x[0], filter(lambda x: any(map(lambda item: True if type(item) == int and item > 0 else False, x[1])), minekonom_result_ilicha.items())))
mineko_result_only_positive_uslugi_list_lenina = list(map(lambda x: x[0], filter(lambda x: any(map(lambda item: True if type(item) == int and item > 0 else False, x[1])), minekonom_result_lenina.items())))
# - - - - - - - -
mineko_set_ilicha = set(mineko_result_only_positive_uslugi_list_ilicha)
mineko_set_lenina = set(mineko_result_only_positive_uslugi_list_lenina)
# - - - - - - - -
set_ilicha = set(list(result_ilicha.keys()))
set_lenina = set(list(result_lenina.keys()))
# - - - - - - - -
ilicha = mineko_set_ilicha & set_ilicha
lenina = mineko_set_lenina & set_lenina
print(f'Совпадений услуг из минэко и остальных отчетов: {len(ilicha)} штук по офису Ильича')
print(f'Совпадений услуг из минэко и остальных отчетов: {len(lenina)} штук по офису Ленина')
print('По Ильича')
for item in ilicha:
    print('Услуга :', item)
    print(f'В обч. отч.: {0 if result_ilicha[item][0] == None else result_ilicha[item][0]} | {0 if result_ilicha[item][2] == None else result_ilicha[item][2]} | {0 if result_ilicha[item][3] == None else result_ilicha[item][3]} | {result_ilicha[item][4]} | {result_ilicha[item][5]}')
    print(f'В минэконом: {minekonom_result_ilicha[item][0]} | {minekonom_result_ilicha[item][2]} | {minekonom_result_ilicha[item][3]} | {minekonom_result_ilicha[item][4]} | {minekonom_result_ilicha[item][5]}')
    print('--------------------------------------------------------')
print()
print()
print()
print('По Ленина')
for item in lenina:
    print('Услуга :', item)
    print(f'В обч. отч.: {0 if result_lenina[item][0] == None else result_lenina[item][0]} | {0 if result_lenina[item][2] == None else result_lenina[item][2]} | {0 if result_lenina[item][3] == None else result_lenina[item][3]} | {result_lenina[item][4]} | {result_lenina[item][5]}')
    print(f'В минэконом: {minekonom_result_lenina[item][0]} | {minekonom_result_lenina[item][2]} | {minekonom_result_lenina[item][3]} | {minekonom_result_lenina[item][4]} | {minekonom_result_lenina[item][5]}')
    print('--------------------------------------------------------')
'''